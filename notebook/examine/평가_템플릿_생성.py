#!/usr/bin/env python3
"""
KLSBench 전문가 평가를 위한 Excel 템플릿 생성 스크립트

사용법:
    python3 평가_템플릿_생성.py --input expert_evaluation_sample.json --output 전문가평가_템플릿.xlsx
"""

import json
import argparse
from pathlib import Path


def create_evaluation_template(input_json, output_xlsx):
    """
    JSON 샘플 데이터를 읽어 Excel 평가 템플릿 생성

    Args:
        input_json: 샘플링된 JSON 파일 경로
        output_xlsx: 출력 Excel 파일 경로
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: pandas와 openpyxl이 필요합니다.")
        print("설치: pip install pandas openpyxl")
        return

    # JSON 데이터 로드
    with open(input_json, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Excel Writer 생성
    writer = pd.ExcelWriter(output_xlsx, engine='openpyxl')

    # 1. Classification 시트
    if 'classification' in data:
        rows = []
        for idx, item in enumerate(data['classification'], 1):
            rows.append({
                '번호': idx,
                'ID': item.get('id', ''),
                '원문': item.get('text', ''),
                '현재_라벨': item.get('label', ''),
                '정확성': '',  # 평가자 입력란
                '올바른_라벨': '',
                '난이도(1-5)': '',
                '의견': ''
            })
        df_class = pd.DataFrame(rows)
        df_class.to_excel(writer, sheet_name='Classification', index=False)

    # 2. Retrieval 시트
    if 'retrieval' in data:
        rows = []
        for idx, item in enumerate(data['retrieval'], 1):
            rows.append({
                '번호': idx,
                'ID': item.get('id', ''),
                'Query': item.get('query', ''),
                'Document': item.get('document', ''),
                '관련성': '',  # 평가자 입력란
                '관련없는_이유': '',
                '난이도(1-5)': '',
                '의견': ''
            })
        df_retr = pd.DataFrame(rows)
        df_retr.to_excel(writer, sheet_name='Retrieval', index=False)

    # 3. Punctuation 시트
    if 'punctuation' in data:
        rows = []
        for idx, item in enumerate(data['punctuation'], 1):
            rows.append({
                '번호': idx,
                'ID': item.get('id', ''),
                '원문(구두점_없음)': item.get('input', ''),
                '정답(구두점_있음)': item.get('output', ''),
                '정확성': '',  # 평가자 입력란
                '수정_제안': '',
                '오류_유형': '',
                '난이도(1-5)': '',
                '의견': ''
            })
        df_punc = pd.DataFrame(rows)
        df_punc.to_excel(writer, sheet_name='Punctuation', index=False)

    # 4. NLI 시트
    if 'nli' in data:
        rows = []
        for idx, item in enumerate(data['nli'], 1):
            rows.append({
                '번호': idx,
                'ID': item.get('id', ''),
                'Premise': item.get('premise', ''),
                'Hypothesis': item.get('hypothesis', ''),
                '현재_라벨': item.get('label', ''),
                '정확성': '',  # 평가자 입력란
                '올바른_라벨': '',
                '판단_근거': '',
                '난이도(1-5)': '',
                '의견': ''
            })
        df_nli = pd.DataFrame(rows)
        df_nli.to_excel(writer, sheet_name='NLI', index=False)

    # 5. Translation 시트
    if 'translation' in data:
        rows = []
        for idx, item in enumerate(data['translation'], 1):
            rows.append({
                '번호': idx,
                'ID': item.get('id', ''),
                '한문_원문': item.get('source', ''),
                '제시된_번역': item.get('target', ''),
                '정확성': '',  # 평가자 입력란
                '수정_제안': '',
                '오역_유형': '',
                '자연스러움': '',
                '난이도(1-5)': '',
                '의견': ''
            })
        df_trans = pd.DataFrame(rows)
        df_trans.to_excel(writer, sheet_name='Translation', index=False)

    # 6. 전반적 평가 시트
    summary_data = {
        '태스크': ['Classification', 'Retrieval', 'Punctuation', 'NLI', 'Translation'],
        '전반적_품질': ['', '', '', '', ''],
        '개선_필요_사항': ['', '', '', '', '']
    }
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='전반적평가', index=False)

    writer.close()

    # 스타일 적용
    wb = load_workbook(output_xlsx)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 헤더 스타일
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # 최대 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # 행 높이 조정 (원문이 긴 경우)
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 30

    wb.save(output_xlsx)
    print(f"Excel 템플릿 생성 완료: {output_xlsx}")
    print(f"총 시트 수: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"  - {sheet_name}: {sheet.max_row - 1}개 항목")


def main():
    parser = argparse.ArgumentParser(description='KLSBench 전문가 평가 Excel 템플릿 생성')
    parser.add_argument('--input', required=True, help='입력 JSON 파일 (샘플 데이터)')
    parser.add_argument('--output', default='전문가평가_템플릿.xlsx', help='출력 Excel 파일명')

    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"Error: 입력 파일을 찾을 수 없습니다: {args.input}")
        return

    create_evaluation_template(args.input, args.output)


if __name__ == '__main__':
    main()
